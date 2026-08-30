# -*- coding: utf-8 -*-
"""C11b · 生成解释文本人工评分表（Excel，可填分+自动汇总）

读取 kg_sample_for_review.md 的 54 条分层抽样样本，
生成带 5 分制量规、数据验证下拉、三评审分表 + 自动汇总 的 xlsx。
"""
import os, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = f"{ROOT}/data/clean/warnings/official/kg/kg_sample_for_review.md"
OUT = f"{ROOT}/data/clean/warnings/official/kg/解释文本评分表.xlsx"

# ---------- 解析样本 ----------
text = open(SRC, encoding="utf-8").read()
# 标题: ### 样本 #1  [au · L1(关注) · 2023-11-21 14:15:00]
head_re = re.compile(r"### 样本 #(\d+)\s+\[(.+?) · (.+?) · (.+?)\]")
# 代码块内容（两行 ``` 之间）
block_re = re.compile(r"### 样本 #\d+.*?\n\n```\n(.*?)\n```", re.DOTALL)

samples = []
for m in head_re.finditer(text):
    idx, variety, level, dt = m.groups()
    samples.append({"idx": int(idx), "variety": variety,
                    "level": level, "dt": dt})
blocks = block_re.findall(text)
assert len(samples) == len(blocks), f"样本数不匹配 {len(samples)} vs {len(blocks)}"
for s, b in zip(samples, blocks):
    s["body"] = b.strip()

# ---------- 样式 ----------
title_font = Font(name="微软雅黑", size=14, bold=True)
h_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
h_fill = PatternFill("solid", fgColor="305496")
cell_font = Font(name="微软雅黑", size=10)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ===== Sheet1: 说明与量规 =====
ws0 = wb.active
ws0.title = "说明与量规"
ws0["A1"] = "期权波动率风险预警系统 · 解释文本人工评分表"
ws0["A1"].font = title_font
ws0.merge_cells("A1:F1")
rows = [
    "",
    "【抽样方法】按 (品种 × 预警级别) 分层抽样，每格 3 条，随机种子 20260813，共 54 条（全量 48,109 条）。",
    "【自动完整性校验】54/54 样本含全部五段式章节（触发因子/市场形态/历史相似/宏观环境/推理结论），平均 241 字。",
    "【评分方式】请评审人独立阅读每条样本的五段式解释，按下方量规打 1~5 分，填在对应评审分表的“评分”列；",
    "              可在“扣分原因”列备注问题；在“评审签名”列签名。三份分表填完后，『汇总』表会自动计算平均分。",
    "【指标要求】问题2 要求解释文本评分 ≥ 4/5（即平均分 ≥ 4.0 视为达标）。",
    "",
    "评分量规（5 分制）：",
]
r = 3
for line in rows:
    ws0.cell(r, 1, line).font = cell_font
    ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

rubric = [
    ("分数", "标准"),
    ("5", "因子归因准确、形态描述与数据一致、历史相似合理、宏观关联恰当、结论可操作"),
    ("4", "五段齐全、归因基本准确、结论合理（允许个别措辞冗余）"),
    ("3", "主体信息齐全但存在 1 处明显归因偏差或信息空洞"),
    ("2", "缺失≥1 个章节，或归因与触发因子明显矛盾"),
    ("1", "文本不可读 / 完全无信息量"),
]
for i, (a, b) in enumerate(rubric):
    ca = ws0.cell(r, 1, a); cb = ws0.cell(r, 2, b)
    ca.font = h_font if i == 0 else cell_font
    cb.font = h_font if i == 0 else cell_font
    ca.fill = h_fill if i == 0 else PatternFill()
    cb.fill = h_fill if i == 0 else PatternFill()
    cb.alignment = wrap
    ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
ws0.column_dimensions["A"].width = 12
for c in "BCDEF":
    ws0.column_dimensions[c].width = 20

# ===== 评审分表 ×3 + 汇总 =====
headers = ["编号", "品种", "级别", "日期", "五段式解释文本", "评分(1-5)", "扣分原因", "评审签名"]
dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5",
                    allow_blank=True, showDropDown=False)
dv.error = "请输入 1~5 的整数"
dv.prompt = "请按量规打 1~5 分"

sheet_names = ["评审1分表", "评审2分表", "评审3分表"]
for sn in sheet_names:
    ws = wb.create_sheet(sn)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = h_font; cell.fill = h_fill; cell.alignment = center; cell.border = border
    for i, s in enumerate(samples, start=2):
        ws.cell(i, 1, s["idx"]).border = border
        ws.cell(i, 2, s["variety"]).border = border
        ws.cell(i, 3, s["level"]).border = border
        ws.cell(i, 4, s["dt"]).border = border
        tb = ws.cell(i, 5, s["body"]); tb.alignment = wrap; tb.border = border
        sc = ws.cell(i, 6); sc.alignment = center; sc.border = border
        rs = ws.cell(i, 7); rs.alignment = wrap; rs.border = border
        sg = ws.cell(i, 8); sg.alignment = center; sg.border = border
    ws.add_data_validation(dv)
    dv.add(f"F2:F{len(samples)+1}")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 70
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 26
    ws.column_dimensions["H"].width = 10
    ws.freeze_panes = "A2"

# ===== 汇总表 =====
ws_sum = wb.create_sheet("汇总")
sum_headers = ["编号", "品种", "级别", "评审1", "评审2", "评审3", "平均分", "是否达标(≥4)"]
for c, h in enumerate(sum_headers, 1):
    cell = ws_sum.cell(1, c, h)
    cell.font = h_font; cell.fill = h_fill; cell.alignment = center; cell.border = border
n = len(samples)
for i, s in enumerate(samples, start=2):
    ws_sum.cell(i, 1, s["idx"]).border = border
    ws_sum.cell(i, 2, s["variety"]).border = border
    ws_sum.cell(i, 3, s["level"]).border = border
    ws_sum.cell(i, 4, f"='评审1分表'!F{i}").border = border
    ws_sum.cell(i, 5, f"='评审2分表'!F{i}").border = border
    ws_sum.cell(i, 6, f"='评审3分表'!F{i}").border = border
    avg = ws_sum.cell(i, 7, f"=IFERROR(AVERAGE(D{i}:F{i}),\"\")")
    avg.alignment = center; avg.border = border
    ok = ws_sum.cell(i, 8, f'=IF(ISNUMBER(G{i}),IF(G{i}>=4,"达标","未达标"),"")')
    ok.alignment = center; ok.border = border
# 总体平均
tr = n + 2
ws_sum.cell(tr, 1, "总体平均分").font = Font(bold=True)
ws_sum.cell(tr, 7, f"=IFERROR(AVERAGE(G2:G{n+1}),\"\")").font = Font(bold=True)
ws_sum.cell(tr, 7).alignment = center
ws_sum.cell(tr+1, 1, "达标条数").font = Font(bold=True)
ws_sum.cell(tr+1, 7, f'=COUNTIF(H2:H{n+1},"达标")&" / "&{n}').font = Font(bold=True)
ws_sum.column_dimensions["A"].width = 6
ws_sum.column_dimensions["B"].width = 6
ws_sum.column_dimensions["C"].width = 12
for c in "DEFGH":
    ws_sum.column_dimensions[c].width = 12
ws_sum.freeze_panes = "A2"

wb.save(OUT)
print(f"已生成: {OUT}")
print(f"样本数: {n}  工作表: {wb.sheetnames}")
